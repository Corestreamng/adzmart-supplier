from django import forms
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import get_template
from django.shortcuts import reverse
from django.core.mail import EmailMultiAlternatives
import openpyxl, os
from cloudinary.forms import CloudinaryJsFileField
from django.utils.safestring import mark_safe
import logging

from apps.authentication.tokens import generate_account_activation_token
from apps.catalog.models import (
    Unit,
    UnitType,
    RadioUnit,
    TVUnit,
    BillboardImage,
    CinemaUnit,
    SpecialOffers,
    PrintUnit,
)
from apps.catalog.utilities import validate_file_extension, get_file_headers, XLHEADERS
from apps.authentication.models import User

log = logging.getLogger(__name__)


class UnitFormAdmin(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.fields != {}:
            self.fields["longitude"].required = True
            self.fields["latitude"].required = True
            self.fields["name"].required = True

    class Meta:
        model = Unit
        exclude = ["created_by", "deleted_at", "user"]


class ExcelUploadForm(forms.Form):
    file = forms.FileField(
        label="Upload Catalog Units",
        validators=[validate_file_extension],
        widget=forms.FileInput(attrs={"class": "form-control"}),
        help_text="-Only excel spreadsheet(.xls, .xlsx) containing the catalog units is accepted.",
    )

    def upload_errors(self, absent_cols_conv, unknown_cols_conv):
        err_msg = ""
        if len(absent_cols_conv):
            err_msg += "The following column(s) are missing: %s." % (
                ", ".join(absent_cols_conv)
            )

        if len(unknown_cols_conv):
            err_msg += os.linesep
            err_msg += "Unknown column(s) in file: %s." % (", ".join(unknown_cols_conv))

        return err_msg

    def save(self, request, form, **kwargs):
        file = self.cleaned_data["file"]
        wb = openpyxl.load_workbook(filename=file, read_only=False)
        tab_names = wb.get_sheet_names()

        for tab_name in tab_names:
            ws = wb[tab_name]
            err_msg = None
            if kwargs["action"] == "catalog:upload_billboard_catalog":
                (
                    headers_row,
                    headers,
                    absent_cols_conv_billboard,
                    unknown_cols_conv_billboard,
                    _,
                    _,
                ) = get_file_headers(ws, action=kwargs["action"])
                err_msg = self.upload_errors(
                    absent_cols_conv_billboard, unknown_cols_conv_billboard
                )

                # Save Billboard entries to database
                unit_types = []
                unit_obj = None

                # save catalog units to db while checking for unique constraint
                for i, row in enumerate(ws.rows):
                    if i <= headers_row:  # Skip to actual entries
                        continue

                    unit_types.append(row[headers[XLHEADERS.UNIT_TYPE]].value)
                    # remove duplicate unit types
                    unit_types_unique = set(unit_types)

                    # save unit type as unique entity
                    for _unit in unit_types_unique:
                        if not UnitType.objects.filter(name=_unit).exists():
                            unit_type_obj = UnitType.objects.create(name=_unit)
                            unit_type_obj.save()
                    try:
                        if row[headers[XLHEADERS.UNIT_TYPE]].value != None:
                            unit_obj = Unit.objects.create(
                                user=request.user.supplier.owner,
                                name=row[headers[XLHEADERS.NAME]].value,
                                supplier=row[headers[XLHEADERS.SUPPLIER]].value,
                                unit_type=UnitType.objects.filter(
                                    name=row[headers[XLHEADERS.UNIT_TYPE]].value
                                ).first(),
                                display_name=row[headers[XLHEADERS.DISPLAY_NAME]].value,
                                reference_id=row[headers[XLHEADERS.REFERENCE_ID]].value,
                                billboard_id=row[headers[XLHEADERS.BILLBOARD_ID]].value,
                                latitude=row[headers[XLHEADERS.LATITUDE]].value,
                                longitude=row[headers[XLHEADERS.LONGITUDE]].value,
                                district=row[headers[XLHEADERS.DISTRICT]].value,
                                state=row[headers[XLHEADERS.STATE]].value,
                                postal_code=row[headers[XLHEADERS.POSTAL_CODE]].value,
                                country=row[headers[XLHEADERS.COUNTRY]].value,
                                facing=row[headers[XLHEADERS.FACING]].value,
                                description=row[headers[XLHEADERS.DESCRIPTION]].value,
                            )
                    except Exception as e:
                        messages.error(
                            request,
                            f"Error: We could not upload your inventory as the Billboard Unit(s) already exists!",
                        )
                        log.error(f"Error creating inventory - {e} ", exc_info=1)
                        break

                if unit_obj != None:
                    messages.success(request, message="Inventory upload successful")
                    return unit_obj

            elif kwargs["action"] == "catalog:upload_radio_catalog":
                ws = wb[tab_name]
                (
                    headers_row,
                    headers,
                    _,
                    _,
                    absent_cols_conv_imod,
                    unknown_cols_conv_imod,
                ) = get_file_headers(ws, action=kwargs["action"])

                err_msg = self.upload_errors(
                    absent_cols_conv_imod, unknown_cols_conv_imod
                )

                # Save Radio entries to database
                for i, row in enumerate(ws.rows):
                    if i <= headers_row:  # Skip to actual entries
                        continue

                    # save imod catalog to db
                    # possible check if entry already exists? what if suppliers have similar inventory???
                    # we do not want to populate db with duplicate rate entries
                    exisiting_inventory = RadioUnit.objects.filter(
                        user=request.user.supplier.owner,
                        Mp_Code=row[headers[XLHEADERS.MP_CODE]].value,
                        Vendor_Name=row[headers[XLHEADERS.VENDOR_NAME]].value,
                        Corporate_Name=row[headers[XLHEADERS.CORPORATE_NAME]].value,
                        Station_Name=row[headers[XLHEADERS.STATION_NAME]].value,
                        State=row[headers[XLHEADERS.STATE_NAME]].value,
                        Media_Type=row[headers[XLHEADERS.MEDIA_TYPE]].value,
                        Rate_Desc=row[headers[XLHEADERS.RATE_DESC]].value,
                        Time=row[headers[XLHEADERS.TIME]].value,
                        Duration=row[headers[XLHEADERS.DURATION]].value,
                        Card_Rate=row[headers[XLHEADERS.CARD_RATE]].value,
                        Nego_Rate=row[headers[XLHEADERS.NEGO_RATE]].value,
                        Card_VD=row[headers[XLHEADERS.CARD_VD]].value,
                        Nego_VD=row[headers[XLHEADERS.NEGO_VD]].value,
                        Card_SC=row[headers[XLHEADERS.CARD_SC]].value,
                        Nego_SC=row[headers[XLHEADERS.NEGO_SC]].value,
                        Add_VD=row[headers[XLHEADERS.Add_VD]].value,
                        SP_Disc=row[headers[XLHEADERS.SP_DISC]].value,
                        Agency=row[headers[XLHEADERS.AGENCY]].value,
                        VAT=row[headers[XLHEADERS.VAT]].value,
                    ).exists()

                    if not exisiting_inventory:
                        RadioUnit.objects.create(
                            user=request.user.supplier.owner,
                            Mp_Code=row[headers[XLHEADERS.MP_CODE]].value,
                            Vendor_Name=row[headers[XLHEADERS.VENDOR_NAME]].value,
                            Corporate_Name=row[headers[XLHEADERS.CORPORATE_NAME]].value,
                            Station_Name=row[headers[XLHEADERS.STATION_NAME]].value,
                            State=row[headers[XLHEADERS.STATE_NAME]].value,
                            Media_Type=row[headers[XLHEADERS.MEDIA_TYPE]].value,
                            Rate_Desc=row[headers[XLHEADERS.RATE_DESC]].value,
                            Time=row[headers[XLHEADERS.TIME]].value,
                            Duration=row[headers[XLHEADERS.DURATION]].value,
                            Card_Rate=row[headers[XLHEADERS.CARD_RATE]].value,
                            Nego_Rate=row[headers[XLHEADERS.NEGO_RATE]].value,
                            Card_VD=row[headers[XLHEADERS.CARD_VD]].value,
                            Nego_VD=row[headers[XLHEADERS.NEGO_VD]].value,
                            Card_SC=row[headers[XLHEADERS.CARD_SC]].value,
                            Nego_SC=row[headers[XLHEADERS.NEGO_SC]].value,
                            Add_VD=row[headers[XLHEADERS.Add_VD]].value,
                            SP_Disc=row[headers[XLHEADERS.SP_DISC]].value,
                            Agency=row[headers[XLHEADERS.AGENCY]].value,
                            VAT=row[headers[XLHEADERS.VAT]].value,
                        )

                if exisiting_inventory:
                    messages.error(
                        request,
                        "Error: We could not upload your inventory as the Radio Unit(s) already exists!",
                    )
                else:
                    messages.success(
                        request, message="Radio Inventory upload successful"
                    )
            else:
                ws = wb[tab_name]
                (
                    headers_row,
                    headers,
                    _,
                    _,
                    absent_cols_conv_imod,
                    unknown_cols_conv_imod,
                ) = get_file_headers(ws, action=kwargs["action"])

                err_msg = self.upload_errors(
                    absent_cols_conv_imod, unknown_cols_conv_imod
                )
                # Save TV entries to database
                for i, row in enumerate(ws.rows):
                    if i <= headers_row:  # Skip to actual entries
                        continue

                    # save imod catalog to db
                    # possible check if entry already exists? what if suppliers have similar inventory???
                    # we do not want to populate db with duplicate rate entries
                    exisiting_inventory = TVUnit.objects.filter(
                        user=request.user.supplier.owner,
                        Mp_Code=row[headers[XLHEADERS.MP_CODE]].value,
                        Vendor_Name=row[headers[XLHEADERS.VENDOR_NAME]].value,
                        Corporate_Name=row[headers[XLHEADERS.CORPORATE_NAME]].value,
                        Station_Name=row[headers[XLHEADERS.STATION_NAME]].value,
                        State=row[headers[XLHEADERS.STATE_NAME]].value,
                        Media_Type=row[headers[XLHEADERS.MEDIA_TYPE]].value,
                        Rate_Desc=row[headers[XLHEADERS.RATE_DESC]].value,
                        Time=row[headers[XLHEADERS.TIME]].value,
                        Duration=row[headers[XLHEADERS.DURATION]].value,
                        Card_Rate=row[headers[XLHEADERS.CARD_RATE]].value,
                        Nego_Rate=row[headers[XLHEADERS.NEGO_RATE]].value,
                        Card_VD=row[headers[XLHEADERS.CARD_VD]].value,
                        Nego_VD=row[headers[XLHEADERS.NEGO_VD]].value,
                        Card_SC=row[headers[XLHEADERS.CARD_SC]].value,
                        Nego_SC=row[headers[XLHEADERS.NEGO_SC]].value,
                        Add_VD=row[headers[XLHEADERS.Add_VD]].value,
                        SP_Disc=row[headers[XLHEADERS.SP_DISC]].value,
                        Agency=row[headers[XLHEADERS.AGENCY]].value,
                        VAT=row[headers[XLHEADERS.VAT]].value,
                    ).exists()

                    if not exisiting_inventory:
                        TVUnit.objects.create(
                            user=request.user.supplier.owner,
                            Mp_Code=row[headers[XLHEADERS.MP_CODE]].value,
                            Vendor_Name=row[headers[XLHEADERS.VENDOR_NAME]].value,
                            Corporate_Name=row[headers[XLHEADERS.CORPORATE_NAME]].value,
                            Station_Name=row[headers[XLHEADERS.STATION_NAME]].value,
                            State=row[headers[XLHEADERS.STATE_NAME]].value,
                            Media_Type=row[headers[XLHEADERS.MEDIA_TYPE]].value,
                            Rate_Desc=row[headers[XLHEADERS.RATE_DESC]].value,
                            Time=row[headers[XLHEADERS.TIME]].value,
                            Duration=row[headers[XLHEADERS.DURATION]].value,
                            Card_Rate=row[headers[XLHEADERS.CARD_RATE]].value,
                            Nego_Rate=row[headers[XLHEADERS.NEGO_RATE]].value,
                            Card_VD=row[headers[XLHEADERS.CARD_VD]].value,
                            Nego_VD=row[headers[XLHEADERS.NEGO_VD]].value,
                            Card_SC=row[headers[XLHEADERS.CARD_SC]].value,
                            Nego_SC=row[headers[XLHEADERS.NEGO_SC]].value,
                            Add_VD=row[headers[XLHEADERS.Add_VD]].value,
                            SP_Disc=row[headers[XLHEADERS.SP_DISC]].value,
                            Agency=row[headers[XLHEADERS.AGENCY]].value,
                            VAT=row[headers[XLHEADERS.VAT]].value,
                        )
                if exisiting_inventory:
                    messages.error(
                        request,
                        "Error: We could not upload your inventory as the TV Unit(s) already exists!",
                    )
                else:
                    messages.success(request, message="TV Inventory upload successful")

            if err_msg:
                messages.error(request, err_msg)


class UnitForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.fields != {}:
            self.fields["longitude"].required = True
            self.fields["latitude"].required = True
            self.fields["name"].required = True
            self.fields["description"].required = True
            self.fields["district"].required = True

    class Meta:
        model = Unit
        fields = [
            "name",
            "display_name",
            "unit_type",
            "reference_id",
            "billboard_id",
            "description",
            "facing",
            "longitude",
            "latitude",
            "district",
            "postal_code",
        ]


class BillboardImageForm(forms.ModelForm):
    image = CloudinaryJsFileField(
        attrs={"multiple": True},
        help_text=mark_safe(
            '<span class="help-text-ul">'
            '<span class="help-text-li">Select all images to upload.</span>'
            '<span class="help-text-li">Click done after uploading all billboard images.</span>'
            "</span>"
        ),
    )

    class Meta:
        model = BillboardImage
        fields = ["image"]


class RadioUnitForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.fields != {}:
            self.fields["Mp_Code"].required = True
            self.fields["Vendor_Name"].required = True
            self.fields["Corporate_Name"].required = True
            self.fields["Station_Name"].required = True
            self.fields["State"].required = True
            self.fields["Media_Type"].required = True
            self.fields["Time"].required = True
            self.fields["Rate_Desc"].required = True
            self.fields["Duration"].required = True
            self.fields["Card_Rate"].required = True
            self.fields["Nego_Rate"].required = True
            self.fields["Card_VD"].required = True
            self.fields["Card_SC"].required = True
            self.fields["Card_SC"].required = True
            self.fields["Nego_SC"].required = True
            self.fields["Agency"].required = True
            self.fields["VAT"].required = True

    class Meta:
        model = RadioUnit
        fields = [
            "Mp_Code",
            "Vendor_Name",
            "Corporate_Name",
            "Station_Name",
            "State",
            "Media_Type",
            "Time",
            "Rate_Desc",
            "Duration",
            "Card_Rate",
            "Nego_Rate",
            "Card_VD",
            "Nego_VD",
            "Card_SC",
            "Nego_SC",
            "Add_VD",
            "SP_Disc",
            "Agency",
            "VAT",
        ]


class AddStaffForm(forms.ModelForm):
    """
    Form for a supplier to create a staff user
    """

    email = forms.EmailField(label="Email", required=True)
    first_name = forms.CharField(label="First Name", required=True)
    last_name = forms.CharField(label="Last Name", required=True)
    phone_no = forms.CharField(label="Phone", required=True)

    class Meta:
        model = User
        fields = ("email", "first_name", "last_name", "phone_no")

    def clean_email(self):
        email = self.cleaned_data.get("email").lower()
        if User.objects.filter(email=email).exists():
            raise ValidationError("Email already exists")
        return email

    def send_activation_email(self, request, user):
        context = {
            "user": user,
            "email_activation_link": request.build_absolute_uri(
                reverse(
                    "password_reset_confirm",
                    kwargs={
                        "uidb64": urlsafe_base64_encode(force_bytes(user.pk)),
                        "token": generate_account_activation_token.make_token(user),
                    },
                )
            ),
        }
        html_content = get_template("staff-user-verification.html").render(
            {"context": context}
        )
        text_content = get_template("staff-user-verification.txt").render(
            {"context": context}
        )
        subject = "Adzmart Account Verification"

        try:
            email = EmailMultiAlternatives(
                subject=subject, body=text_content, from_email=None, to=[user.email]
            )
            email.attach_alternative(html_content, "text/html")
            EmailThread(email).start()
            messages.success(
                request, message="User added and email notification sent successfully!"
            )
        except Exception as e:
            messages.error(request, e)


class TVUnitForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.fields != {}:
            self.fields["Mp_Code"].required = True
            self.fields["Vendor_Name"].required = True
            self.fields["Corporate_Name"].required = True
            self.fields["Station_Name"].required = True
            self.fields["State"].required = True
            self.fields["Media_Type"].required = True
            self.fields["Time"].required = True
            self.fields["Rate_Desc"].required = True
            self.fields["Duration"].required = True
            self.fields["Card_Rate"].required = True
            self.fields["Nego_Rate"].required = True
            self.fields["Card_VD"].required = True
            self.fields["Card_SC"].required = True
            self.fields["Card_SC"].required = True
            self.fields["Nego_SC"].required = True
            self.fields["Agency"].required = True
            self.fields["VAT"].required = True

    class Meta:
        model = TVUnit
        fields = [
            "Mp_Code",
            "Vendor_Name",
            "Corporate_Name",
            "Station_Name",
            "State",
            "Media_Type",
            "Time",
            "Rate_Desc",
            "Duration",
            "Card_Rate",
            "Nego_Rate",
            "Card_VD",
            "Nego_VD",
            "Card_SC",
            "Nego_SC",
            "Add_VD",
            "SP_Disc",
            "Agency",
            "VAT",
        ]


class CinemaUnitForm(forms.ModelForm):
    cinema = forms.CharField(
        label="Cinema", required=True, help_text="Name of Cinema. e.g SilverBird"
    )
    location = forms.CharField(
        label="Location", required=True, help_text="Location of Cinema e.g Ikeja Mall"
    )
    rate_per_spot = forms.IntegerField(label="Rate per Spot", required=True)
    state = forms.CharField(
        label="State",
        required=True,
        help_text="State where cinema is located. e.g Lagos",
    )

    class Meta:
        model = CinemaUnit
        fields = ["cinema", "location", "rate_per_spot", "state"]


class CinemaUnitExcelUploadForm(forms.Form):
    file = forms.FileField(
        label="Upload Cinema Units",
        validators=[validate_file_extension],
        widget=forms.FileInput(attrs={"class": "form-control"}),
        help_text="-Only excel spreadsheet(.xls, .xlsx) containing the cinema units is accepted.",
    )

    class CINEMAHEADERS:
        CINEMA = ("cinema",)
        LOCATION = ("location",)
        RATE_PER_SPOT = ("rate_per_spot",)
        STATE_CINEMA = "state"

        cinema_choices = [CINEMA, LOCATION, RATE_PER_SPOT, STATE_CINEMA]

    def get_file_headers(self, ws, action):
        headers_row = None
        headers = {}

        # Get headers row
        for row in range(ws.max_row + 1):
            # consider empty rows in worksheet and get the fist instance of the headers
            cell = ws["A"][row].value

            # select header choices based on url
            header_choices = None
            if action == "catalog:upload_cinema_catalog":
                header_choices = list(map("".join, self.CINEMAHEADERS().cinema_choices))

            if isinstance(cell, str) and cell in header_choices:
                headers_row = row
                break

        # capture rows with no headers
        if headers_row is None:
            return None, None

        # remember position of headers
        file_header = []
        for i in range(ws.max_column):
            column = chr(i + 65)  # go through columns using ASCII('A'-'Z')
            header = ws[column][headers_row].value
            file_header.append(header)
            if header is None:
                break

            # check if header is in XLHEADERS cinema
            for choice in self.CINEMAHEADERS().cinema_choices:
                if header in choice:
                    headers[choice] = i
                    break

        # capture absent and unkown headers in cinema upload
        absent_cols_cinema = set(self.CINEMAHEADERS().cinema_choices).difference(
            headers
        )
        unknown_cols_cinema = set(file_header).difference(list(map("".join, headers)))

        # convert list of tuples to list of strings in cinema upload
        absent_cols_conv_cinema = list(map("".join, absent_cols_cinema))
        unknown_cols_conv_cinema = list(map("".join, unknown_cols_cinema))

        return (headers_row, headers, absent_cols_conv_cinema, unknown_cols_conv_cinema)

    def upload_errors(self, absent_cols_conv, unknown_cols_conv):
        err_msg = ""
        if len(absent_cols_conv):
            err_msg += "The following column(s) are missing: %s." % (
                ", ".join(absent_cols_conv)
            )

        if len(unknown_cols_conv):
            err_msg += os.linesep
            err_msg += "Unknown column(s) in file: %s." % (", ".join(unknown_cols_conv))

        return err_msg

    def save(self, request, form, **kwargs):
        file = self.cleaned_data["file"]
        wb = openpyxl.load_workbook(filename=file, read_only=False)
        tab_names = wb.get_sheet_names()

        for tab_name in tab_names:
            ws = wb[tab_name]
            err_msg = None
            ws = wb[tab_name]
            (
                headers_row,
                headers,
                absent_cols_conv_cinema,
                unknown_cols_conv_cinema,
            ) = self.get_file_headers(ws, action=kwargs["action"])

            err_msg = self.upload_errors(
                absent_cols_conv_cinema, unknown_cols_conv_cinema
            )
            # Save Cinema entries to database
            for i, row in enumerate(ws.rows):
                if i <= headers_row:  # Skip to actual entries
                    continue
                exisiting_inventory = CinemaUnit.objects.filter(
                    user=request.user.supplier.owner,
                    cinema=row[headers[self.CINEMAHEADERS().CINEMA]].value,
                    location=row[headers[self.CINEMAHEADERS().LOCATION]].value,
                    rate_per_spot=row[
                        headers[self.CINEMAHEADERS().RATE_PER_SPOT]
                    ].value,
                    state=row[headers[self.CINEMAHEADERS().STATE_CINEMA]].value,
                ).exists()

                if not exisiting_inventory:
                    CinemaUnit.objects.create(
                        user=request.user.supplier.owner,
                        cinema=row[headers[self.CINEMAHEADERS().CINEMA]].value,
                        location=row[headers[self.CINEMAHEADERS().LOCATION]].value,
                        rate_per_spot=row[
                            headers[self.CINEMAHEADERS().RATE_PER_SPOT]
                        ].value,
                        state=row[headers[self.CINEMAHEADERS().STATE_CINEMA]].value,
                    )
            if exisiting_inventory:
                messages.error(
                    request,
                    "Error: We could not upload your inventory as the Cinema Unit(s) already exists!",
                )
            else:
                messages.success(request, message="Cinema Inventory upload successful")



class SpecialOffersForm(forms.ModelForm):
    class Meta:
        model = SpecialOffers
        exclude = ("user",)

class PrintUnitForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super(PrintUnitForm, self).__init__(*args, **kwargs)

        for key in self.fields:
            self.fields[key].required = True

    position = forms.CharField(
        label="Position", help_text="e.g Political, Loose Inserts."
    )
    size = forms.CharField(
        label="Coverage", help_text='Page size e.g Full Page, 10" X 6COL.'
    )
    vat = forms.CharField(label="VAT")
    agency_discount = forms.CharField(label="Agency Discount")

    class Meta:
        model = PrintUnit
        fields = [
            "coverage",
            "publisher",
            "title",
            "type",
            "size",
            "position",
            "rate",
            "agency_discount",
            "amount",
            "vat",
            "total",
        ]


class PrintUnitExcelUploadForm(forms.Form):
    file = forms.FileField(
        label="Upload Print Units",
        validators=[validate_file_extension],
        widget=forms.FileInput(attrs={"class": "form-control"}),
        help_text="-Only excel spreadsheet(.xls, .xlsx) containing the print units is accepted.",
    )

    class PRINTHEADERS:
        COVERAGE = ("Coverage",)
        PUBLISHER = ("Publisher",)
        TITLE = ("Title",)
        TYPE = ("Type",)
        SIZE = ("Size",)
        POSITION = ("Position",)
        RATE = ("Rate",)
        AGENCY_DISCOUNT = ("Agency Discount",)
        AMOUNT = ("Amount",)
        VAT = ("Vat",)
        TOTAL = ("Total",)

        print_choices = [
            COVERAGE,
            PUBLISHER,
            TITLE,
            TYPE,
            SIZE,
            POSITION,
            RATE,
            AGENCY_DISCOUNT,
            AMOUNT,
            VAT,
            TOTAL,
        ]

    def get_file_headers(self, ws, action):
        headers_row = None
        headers = {}

        # Get headers row
        for row in range(ws.max_row + 1):
            # consider empty rows in worksheet and get the fist instance of the headers
            cell = ws["A"][row].value

            # select header choices based on url
            header_choices = None
            if action == "catalog:upload_print_catalog":
                header_choices = list(map("".join, self.PRINTHEADERS().print_choices))

            if isinstance(cell, str) and cell in header_choices:
                headers_row = row
                break

        # capture rows with no headers
        if headers_row is None:
            return None, None

        # remember position of headers
        file_header = []
        for i in range(ws.max_column):
            column = chr(i + 65)  # go through columns using ASCII('A'-'Z')
            header = ws[column][headers_row].value
            file_header.append(header)
            if header is None:
                break

            # check if header is in XLHEADERS cinema
            for choice in self.PRINTHEADERS().print_choices:
                if header in choice:
                    headers[choice] = i
                    break

        # capture absent and unkown headers in cinema upload
        absent_cols_print = set(self.PRINTHEADERS().print_choices).difference(headers)
        unknown_cols_print = set(file_header).difference(list(map("".join, headers)))

        # convert list of tuples to list of strings in cinema upload
        absent_cols_conv_print = list(map("".join, absent_cols_print))
        unknown_cols_conv_print = list(map("".join, unknown_cols_print))

        return (headers_row, headers, absent_cols_conv_print, unknown_cols_conv_print)

    def upload_errors(self, absent_cols_conv, unknown_cols_conv):
        err_msg = ""
        if len(absent_cols_conv):
            err_msg += "The following column(s) are missing: %s." % (
                ", ".join(absent_cols_conv)
            )

        if len(unknown_cols_conv):
            err_msg += os.linesep
            err_msg += "Unknown column(s) in file: %s." % (", ".join(unknown_cols_conv))

        return err_msg

    def save(self, request, form, **kwargs):
        file = self.cleaned_data["file"]
        wb = openpyxl.load_workbook(filename=file, read_only=False)
        tab_names = wb.get_sheet_names()

        for tab_name in tab_names:
            ws = wb[tab_name]
            err_msg = None
            (
                headers_row,
                headers,
                absent_cols_conv_print,
                unknown_cols_conv_print,
            ) = self.get_file_headers(ws, action=kwargs["action"])

            err_msg = self.upload_errors(
                absent_cols_conv_print, unknown_cols_conv_print
            )
            # Save Print entries to database
            for i, row in enumerate(ws.rows):
                if i <= headers_row:  # Skip to actual entries
                    continue
                exisiting_inventory = PrintUnit.objects.filter(
                    user=request.user.supplier.owner,
                    coverage=row[headers[self.PRINTHEADERS().COVERAGE]].value,
                    publisher=row[headers[self.PRINTHEADERS().PUBLISHER]].value,
                    title=row[headers[self.PRINTHEADERS().TITLE]].value,
                    type=row[headers[self.PRINTHEADERS().TYPE]].value,
                    size=row[headers[self.PRINTHEADERS().SIZE]].value,
                    position=row[headers[self.PRINTHEADERS().POSITION]].value,
                    rate=row[headers[self.PRINTHEADERS().RATE]].value,
                    agency_discount=row[
                        headers[self.PRINTHEADERS().AGENCY_DISCOUNT]
                    ].value,
                    amount=row[headers[self.PRINTHEADERS().AMOUNT]].value,
                    vat=row[headers[self.PRINTHEADERS().VAT]].value,
                    total=row[headers[self.PRINTHEADERS().TOTAL]].value,
                ).exists()

                if not exisiting_inventory:
                    PrintUnit.objects.create(
                        user=request.user.supplier.owner,
                        coverage=row[headers[self.PRINTHEADERS().COVERAGE]].value,
                        publisher=row[headers[self.PRINTHEADERS().PUBLISHER]].value,
                        title=row[headers[self.PRINTHEADERS().TITLE]].value,
                        type=row[headers[self.PRINTHEADERS().TYPE]].value,
                        size=row[headers[self.PRINTHEADERS().SIZE]].value,
                        position=row[headers[self.PRINTHEADERS().POSITION]].value,
                        rate=row[headers[self.PRINTHEADERS().RATE]].value,
                        agency_discount=row[
                            headers[self.PRINTHEADERS().AGENCY_DISCOUNT]
                        ].value,
                        amount=row[headers[self.PRINTHEADERS().AMOUNT]].value,
                        vat=row[headers[self.PRINTHEADERS().VAT]].value,
                        total=row[headers[self.PRINTHEADERS().TOTAL]].value,
                    )
            if exisiting_inventory:
                messages.error(
                    request,
                    "Error: We could not upload your inventory as the Print Unit(s) already exists!",
                )
            else:
                messages.success(request, message="Print Inventory upload successful")

